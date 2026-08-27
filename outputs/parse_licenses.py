# -*- coding: utf-8 -*-
"""解析「各店铺许可证收集(2).xlsx」：建立 店铺行 → 证件类型 → 图片文件 的完整映射"""
import zipfile, re, sys, json
from openpyxl import load_workbook

XLSX = r"C:/Users/Administrator/Desktop/各店铺许可证收集(2).xlsx"
z = zipfile.ZipFile(XLSX)

# ---------- 1) cellimages.xml: DISPIMG ID -> rId + descr ----------
cell_xml = z.read('xl/cellimages.xml').decode('utf-8')
# 结构：<xdr:cNvPr id="2" name="ID_EA44..." descr="铜陵..."/>...<a:blip r:embed="rId1"/>
# 注意：生产许可证的 cNvPr 没有 descr 属性（descr 可选）
id2rid = {}
id2desc = {}
for m in re.finditer(r'name="(ID_[0-9A-Fa-f]+)"(?:\s+descr="([^"]*)")?[\s\S]*?r:embed="(rId\d+)"', cell_xml):
    id2rid[m.group(1)] = m.group(3)
    id2desc[m.group(1)] = m.group(2) or ''
print(f"[cellimages] 发现 {len(id2rid)} 个图片 ID 映射")

# ---------- 2) cellimages.xml.rels: rId -> media/imageN ----------
rels_xml = z.read('xl/_rels/cellimages.xml.rels').decode('utf-8')
rid2media = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml))
print(f"[rels] 发现 {len(rid2media)} 个 rId 映射")

# ---------- 3) 读取 Excel 单元格 ----------
wb = load_workbook(XLSX, data_only=False)
ws = wb.active
print(f"[sheet] {ws.title}  行数={ws.max_row} 列数={ws.max_column}")

def dispimg_id(cell):
    v = cell.value
    if isinstance(v, str):
        m = re.search(r'ID_[0-9A-Fa-f]{32}', v)
        if m: return m.group(0)
    return None

rows = []
for r in range(2, ws.max_row + 1):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
    if not any(vals): continue
    platform   = vals[0] or ''
    shop_name  = vals[1] or ''
    shop_short = vals[2] or ''
    company    = vals[3] or ''
    lic_id     = dispimg_id(ws.cell(row=r, column=5))   # 经营许可证
    biz_id     = dispimg_id(ws.cell(row=r, column=6))   # 营业执照
    factory    = vals[7] or ''
    prod_id    = dispimg_id(ws.cell(row=r, column=9))   # 生产许可证
    rows.append({
        'row': r, 'platform': platform, 'shop_name': shop_name, 'shop_short': shop_short,
        'company': company, 'factory': factory,
        'license': lic_id, 'business': biz_id, 'production': prod_id
    })

print(f"[sheet] 有效店铺数据 {len(rows)} 行\n")
for row in rows:
    def media(i):
        if not i: return '-'
        rid = id2rid.get(i, '?')
        return f"{i[:8]}..→{rid}→{rid2media.get(rid, '?')}"
    print(f"行{row['row']:>2} | {row['platform']:<6} | {row['shop_short']:<8} | {row['company']:<6} | "
          f"许可:{media(row['license']):<28} 执照:{media(row['business']):<28} 生产:{media(row['production'])}")

# ---------- 4) 汇总输出 JSON 供上传脚本使用 ----------
out = {'rows': rows, 'id2rid': id2rid, 'rid2media': rid2media}
with open(r"C:/Users/Administrator/WorkBuddy/2026-07-28-10-50-05/outputs/license_map.json", 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=1)
print("\n映射已保存 outputs/license_map.json")
